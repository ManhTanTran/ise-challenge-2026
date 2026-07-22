"""Create an ASCII-safe Kaggle upload bundle and rewrite evidence paths."""

from __future__ import annotations

import ast
import csv
import hashlib
import json
import re
import shutil
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def safe_name(value: str) -> str:
    suffix = Path(value).suffix
    stem = value[: -len(suffix)] if suffix else value
    stem = unicodedata.normalize("NFKD", stem).encode("ascii", "ignore").decode("ascii")
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._") or "file"
    safe_suffix = re.sub(r"[^A-Za-z0-9.]", "", suffix)
    return f"{stem}{safe_suffix}"


def main(target_name: str = "kaggle_safe_upload_bundle") -> None:
    root = Path(__file__).parents[1]
    source = root / "kaggle_all_chunking_bundle"
    target = root / target_name
    if target.exists():
        raise SystemExit(f"Target already exists: {target}; remove/move it before rebuilding.")
    target.mkdir()

    # Copy code/assets, excluding the large source corpus and transient caches.
    excluded = {"data", "benchmark_questions.xlsx", "dataset_manifest.csv"}
    for item in source.iterdir():
        if item.name in excluded or item.name in {"__pycache__"}:
            continue
        destination = target / item.name
        if item.is_dir():
            shutil.copytree(item, destination, ignore=shutil.ignore_patterns(".git", "__pycache__", "*.pyc"))
        else:
            shutil.copy2(item, destination)

    source_data = source / "data" / "text_sources"
    target_data = target / "data" / "text_sources"
    target_data.mkdir(parents=True)
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for path in sorted(p for p in source_data.rglob("*") if p.is_file()):
        original = path.relative_to(source_data).as_posix()
        parts = [safe_name(part) for part in Path(original).parts]
        candidate = Path(*parts).as_posix()
        if candidate in used:
            stem = Path(candidate).stem
            suffix = Path(candidate).suffix
            digest = hashlib.sha256(original.encode()).hexdigest()[:8]
            candidate = str(Path(candidate).with_name(f"{stem}_{digest}{suffix}")).replace("\\", "/")
        used.add(candidate)
        mapping[original] = candidate
        destination = target_data / Path(candidate)
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, destination)

    def rewrite_evidence_path(value: object) -> str:
        expected = str(value).replace("\\", "/")
        if expected in mapping:
            return mapping[expected]
        matches = [safe for original, safe in mapping.items() if original.endswith("/" + expected)]
        if len(matches) == 1:
            return matches[0]
        return expected

    # Rewrite evidence values while preserving every other workbook cell.
    workbook = load_workbook(source / "benchmark_questions.xlsx")
    for sheet in workbook.worksheets:
        header_row = None
        evidence_column = None
        for row in sheet.iter_rows():
            values = [str(cell.value or "").strip().casefold() for cell in row]
            if "evidences" in values:
                header_row = row[0].row
                evidence_column = values.index("evidences") + 1
                break
        if not header_row or not evidence_column:
            continue
        for row in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row, evidence_column)
            raw = cell.value
            if not raw:
                continue
            try:
                values = json.loads(str(raw))
            except Exception:
                try:
                    values = ast.literal_eval(str(raw))
                except Exception:
                    values = [part.strip().strip("\"'") for part in str(raw).strip("[]").split(",") if part.strip()]
            if isinstance(values, list):
                rewritten = [rewrite_evidence_path(value) for value in values]
                cell.value = json.dumps(rewritten, ensure_ascii=False)
    workbook.save(target / "benchmark_questions.xlsx")

    with (target / "dataset_manifest.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["original_path", "safe_path", "bytes", "sha256"])
        writer.writeheader()
        for original, safe in mapping.items():
            path = target_data / safe
            writer.writerow({
                "original_path": original,
                "safe_path": safe,
                "bytes": path.stat().st_size,
                "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            })

    (target / "KAGGLE_UPLOAD_README.txt").write_text(
        "This is the Kaggle-safe ASCII upload bundle. Evidence paths in benchmark_questions.xlsx "
        "are rewritten to the safe names; dataset_manifest.csv maps them back to original names.\n",
        encoding="utf-8",
    )
    print(target)
    print(f"files={len(mapping)}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--target-name", default="kaggle_safe_upload_bundle")
    main(parser.parse_args().target_name)
