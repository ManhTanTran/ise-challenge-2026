"""Repair a Kaggle-safe benchmark workbook using its original-to-safe manifest."""

from __future__ import annotations

import ast
import csv
import json
from pathlib import Path

from openpyxl import load_workbook


def main(dataset_root: Path) -> None:
    with (dataset_root / "dataset_manifest.csv").open(encoding="utf-8-sig", newline="") as handle:
        mapping = {row["original_path"]: row["safe_path"] for row in csv.DictReader(handle)}

    def rewrite(value: object) -> str:
        expected = str(value).replace("\\", "/")
        if expected in mapping:
            return mapping[expected]
        matches = [safe for original, safe in mapping.items() if original.endswith("/" + expected)]
        return matches[0] if len(matches) == 1 else expected

    workbook_path = dataset_root / "benchmark_questions.xlsx"
    workbook = load_workbook(workbook_path)
    for sheet in workbook.worksheets:
        evidence_column = None
        header_row = None
        for row in sheet.iter_rows():
            headers = [str(cell.value or "").strip().casefold() for cell in row]
            if "evidences" in headers:
                evidence_column = headers.index("evidences") + 1
                header_row = row[0].row
                break
        if not evidence_column or not header_row:
            continue
        for row_number in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row_number, evidence_column)
            if not cell.value:
                continue
            try:
                values = json.loads(str(cell.value))
            except Exception:
                try:
                    values = ast.literal_eval(str(cell.value))
                except Exception:
                    values = [part.strip().strip("\"'") for part in str(cell.value).strip("[]").split(",") if part.strip()]
            if isinstance(values, list):
                cell.value = json.dumps([rewrite(value) for value in values], ensure_ascii=False)
    workbook.save(workbook_path)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("dataset_root", type=Path)
    main(parser.parse_args().dataset_root)
